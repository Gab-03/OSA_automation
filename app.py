import pandas as pd
import openpyxl
from datetime import datetime
from docxtpl import DocxTemplate
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import threading
import time
import xlwings as xw
import warnings
warnings.filterwarnings('ignore')

class FileSelectorApp:
    def __init__(self, master):
        self.master = master
        self.master.title("OSA")
        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        self.loading_window = None
        self.progress_var = tk.DoubleVar()
        self.create_widgets()

    def create_widgets(self):
        self.create_file_entry(0, "Store Master", self.file1_path, self.browse_file1)
        self.create_file_entry(1, "EANs", self.file2_path, self.browse_file2)
        tk.Button(self.master, text="Prepare Form", command=self.submit).grid(row=3, column=1, pady=20)

    def create_file_entry(self, row, label_text, var, command):
        tk.Label(self.master, text=label_text).grid(row=row, column=0, pady=10)
        tk.Entry(self.master, textvariable=var, width=50, state='disabled').grid(row=row, column=1, pady=10)
        tk.Button(self.master, text="Browse", command=command).grid(row=row, column=2, pady=10)

    def browse_file1(self):
        self.file1_path.set(filedialog.askopenfilename())

    def browse_file2(self):
        self.file2_path.set(filedialog.askopenfilename())

    def submit(self):
        file1 = self.file1_path.get()
        file2 = self.file2_path.get()

        if file1 and file2:
            try:
                # Show loading screen before reading files
                self.show_loading_screen()

                # Start background task
                threading.Thread(target=self.read_files_and_process, args=(file1,file2)).start()
                messagebox.showinfo("Process Complete!", "file downloaded successfully.")


            except Exception as e:
                # Display an error message if an error occurs
                messagebox.showerror("Error", f"An error occurred: {str(e)}")

        else:
            messagebox.showerror("Error", "Please upload all the required files.")

    def read_files_and_process(self, file1,file2):

        store_master = pd.read_excel(file1, sheet_name='Store Master')

        #Filter Non PS Door
        store_master = store_master[store_master["Type of Door"] != "Non PS Door"]


        #Filter on Channel Group: DT MAG, OG SMKT, ROG SMKT, SMALL FORMATS.
        store_master = store_master[(store_master["Channel Group"] == "DT MAG") | (store_master["Channel Group"] == "OG SMKT") | (store_master["Channel Group"] == "ROG SMKT") | (store_master["Channel Group"] == "SMALL FORMATS")]

        #Filter on Area/Customer Group: exclude NWS and SM
        store_master = store_master[(store_master["Area/Customer Group"] != "SM") & (store_master["Area/Customer Group"] != "NWS SMKT")]


        regions = ["GMA","LUZON", "VISAYAS","MINDANAO"]
        store_types = ["MASS","HEAVY RESELLER","PREMIUM","SPREMIUM","HYBRID"]

        GMA_MASS = []
        GMA_HR = []
        GMA_P = []
        GMA_SP = []
        GMA_H = []

        LUZON_MASS = []
        LUZON_HR = []
        LUZON_P = []
        LUZON_SP = []
        LUZON_H = []

        VISAYAS_MASS = []
        VISAYAS_HR = []
        VISAYAS_P = []
        VISAYAS_SP = []
        VISAYAS_H = []

        MINDANAO_MASS = []
        MINDANAO_HR = []
        MINDANAO_P = []
        MINDANAO_SP = []
        MINDANAO_H = []


        for index,row in store_master.iterrows():
            region = str(row["Region"]) #store type
            ST = str(row["Store Type"])

            if region == "GMA":
                if "MASS" in ST.split() or "MASS-P" in ST.split():
                    GMA_MASS.append(ST)
                elif "HEAVY" in ST.split() and ("RESELLER" in ST.split() or "RESELLER-P" in ST.split())  :
                    GMA_HR.append(ST)
                elif "PREMIUM" in ST.split() or "PREMIUM-P" in ST.split():
                    GMA_P.append(ST)
                elif "SPREMIUM" in ST.split() or "SPREMIUM-P" in ST.split():
                    GMA_SP.append(ST)
                elif "HYBRID" in ST.split() or  "HYBRID-P" in ST.split():
                    GMA_H.append(ST)

            elif region == "LUZON":
                if "MASS" in ST.split() or "MASS-P" in ST.split():
                    LUZON_MASS.append(ST)
                elif "HEAVY" in ST.split() and ("RESELLER" in ST.split() or "RESELLER-P" in ST.split()) :
                    LUZON_HR.append(ST)
                elif "PREMIUM" in ST.split() or "PREMIUM-P" in ST.split():
                    LUZON_P.append(ST)
                elif "SPREMIUM" in ST.split() or "SPREMIUM-P" in ST.split():
                    LUZON_SP.append(ST)
                elif "HYBRID" in ST.split() or  "HYBRID-P" in ST.split():
                    LUZON_H.append(ST)

            elif region == "VISAYAS":
                if "MASS" in ST.split() or "MASS-P" in ST.split():
                    VISAYAS_MASS.append(ST)
                elif "HEAVY" in ST.split() and ("RESELLER" in ST.split() or "RESELLER-P" in ST.split()) :
                    VISAYAS_HR.append(ST)
                elif "PREMIUM" in ST.split() or "PREMIUM-P" in ST.split():
                    VISAYAS_P.append(ST)
                elif "SPREMIUM" in ST.split() or "SPREMIUM-P" in ST.split():
                    VISAYAS_SP.append(ST)
                elif "HYBRID" in ST.split() or  "HYBRID-P" in ST.split():
                    VISAYAS_H.append(ST)

            elif region == "MINDANAO":
                if "MASS" in ST.split() or "MASS-P" in ST.split():
                    MINDANAO_MASS.append(ST)
                elif "HEAVY" in ST.split() and ("RESELLER" in ST.split() or "RESELLER-P" in ST.split())  :
                    MINDANAO_HR.append(ST)
                elif "PREMIUM" in ST.split() or "PREMIUM-P" in ST.split():
                    MINDANAO_P.append(ST)
                elif "SPREMIUM" in ST.split() or "SPREMIUM-P" in ST.split():
                    MINDANAO_SP.append(ST)
                elif "HYBRID" in ST.split() or  "HYBRID-P" in ST.split():
                    MINDANAO_H.append(ST)

        #removing duplicates
        GMA_MASS = list(set(GMA_MASS))
        GMA_HR = list(set(GMA_HR))
        GMA_P = list(set(GMA_P))
        GMA_SP = list(set(GMA_SP))
        GMA_H = list(set(GMA_H))

        LUZON_MASS = list(set(LUZON_MASS))
        LUZON_HR = list(set(LUZON_HR))
        LUZON_P = list(set(LUZON_P))
        LUZON_SP = list(set(LUZON_SP))
        LUZON_H = list(set(LUZON_H))

        VISAYAS_MASS = list(set(VISAYAS_MASS))
        VISAYAS_HR = list(set(VISAYAS_HR))
        VISAYAS_P = list(set(VISAYAS_P))
        VISAYAS_SP = list(set(VISAYAS_SP))
        VISAYAS_H = list(set(VISAYAS_H))

        MINDANAO_MASS = list(set(MINDANAO_MASS))
        MINDANAO_HR = list(set(MINDANAO_HR))
        MINDANAO_P = list(set(MINDANAO_P))
        MINDANAO_SP = list(set(MINDANAO_SP))
        MINDANAO_H = list(set(MINDANAO_H))

        regions = {
            'GMA_MASS': GMA_MASS,
            'GMA_HR': GMA_HR,
            'GMA_P': GMA_P,
            'GMA_SP': GMA_SP,
            'GMA_H': GMA_H,
            'LUZON_MASS': LUZON_MASS,
            'LUZON_HR': LUZON_HR,
            'LUZON_P': LUZON_P,
            'LUZON_SP': LUZON_SP,
            'LUZON_H': LUZON_H,
            'VISAYAS_MASS': VISAYAS_MASS,
            'VISAYAS_HR': VISAYAS_HR,
            'VISAYAS_P': VISAYAS_P,
            'VISAYAS_SP': VISAYAS_SP,
            'VISAYAS_H': VISAYAS_H,
            'MINDANAO_MASS': MINDANAO_MASS,
            'MINDANAO_HR': MINDANAO_HR,
            'MINDANAO_P': MINDANAO_P,
            'MINDANAO_SP': MINDANAO_SP,
            'MINDANAO_H': MINDANAO_H,
        }

        for region_name, region_list in regions.items():
            # print(f"{region_name}: {region_list}")
            add_r = region_name.split('_')[0]
            for x in range(len(region_list)):
                if region_list[x][-1] == "P":
                    region_list[x] = region_list[x].replace("-P","") + " - " + add_r + "-P"
                elif region_list[x][-1] == "H":
                    region_list[x] = region_list[x].replace("-NH","") + " - " + add_r + "-NH"
                else:
                    region_list[x] = region_list[x] + " - " + add_r


        ##### SM
        store_master = pd.read_excel(file1, sheet_name='Store Master')
        store_master = store_master[store_master["Type of Door"] != "Non PS Door"]
        sm = store_master[store_master["Area/Customer Group"] == "SM"]

        SM_MASS = []
        SM_P = []

        for index,row in sm.iterrows():
            ST = str(row["Store Type"])
            if "MASS" in ST.split() or "MASS-P" in ST.split():
                SM_MASS.append(ST)
            elif "PREMIUM" in ST.split() or "PREMIUM-P" in ST.split() or "SPREMIUM" in ST.split() or "SPREMIUM-P" in ST.split():
                SM_P.append(ST)


        SM_MASS = list(set(SM_MASS))
        SM_P = list(set(SM_P))


        ##### COTF
        store_master = pd.read_excel(file1, sheet_name='Store Master')
        store_master = store_master[store_master["Type of Door"] != "Non PS Door"]
        cotf = store_master[store_master["Channel Group"] == "COTF"]

        beauty = []
        go = []
        health = []

        for index, row in cotf.iterrows():
            area = str(row["Area/Customer Group"])
            if area == "BEAUTY":
                beauty.append(row["Store Type"])
            elif area == "GO":
                go.append(row["Store Type"])
            elif area == "HEALTH":
                health.append(row["Store Type"])

        beauty = list(set(beauty))
        go = list(set(go))
        health = list(set(health))


        #### Store Type x EAN Dups

        EAN = pd.read_excel(file1, sheet_name='EAN per Store Type')
        master_wb = openpyxl.load_workbook('EAN_Input.xlsx')
        master_sheet = master_wb['Sheet1']

        ##############################################
        ### MASS
        ##############################################

        #GMA 
        column_values = [cell.value for cell in master_sheet['B']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        GMA_MASS_EAN = column_values[3:last_row_master]

        #LUZON 
        column_values = [cell.value for cell in master_sheet['C']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        LUZON_MASS_EAN = column_values[3:last_row_master]

        #VISAYAS
        column_values = [cell.value for cell in master_sheet['D']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        VISAYAS_MASS_EAN = column_values[3:last_row_master]

        #MINDANAO
        column_values = [cell.value for cell in master_sheet['E']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        MINDANAO_MASS_EAN = column_values[3:last_row_master]

        ##############################################
        ### PREMIUM
        ##############################################

        #GMA 
        column_values = [cell.value for cell in master_sheet['H']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        GMA_PREMIUM_EAN = column_values[3:last_row_master]

        #LUZON 
        column_values = [cell.value for cell in master_sheet['I']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        LUZON_PREMIUM_EAN = column_values[3:last_row_master]

        #VISAYAS
        column_values = [cell.value for cell in master_sheet['J']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        VISAYAS_PREMIUM_EAN = column_values[3:last_row_master]

        #MINDANAO
        column_values = [cell.value for cell in master_sheet['K']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        MINDANAO_PREMIUM_EAN = column_values[3:last_row_master]

        ##############################################
        ### SPREMIUM
        ##############################################

        #GMA 
        column_values = [cell.value for cell in master_sheet['N']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        GMA_SPREMIUM_EAN = column_values[3:last_row_master]

        #LUZON 
        column_values = [cell.value for cell in master_sheet['O']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        LUZON_SPREMIUM_EAN = column_values[3:last_row_master]

        #VISAYAS
        column_values = [cell.value for cell in master_sheet['P']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        VISAYAS_SPREMIUM_EAN = column_values[3:last_row_master]

        #MINDANAO
        column_values = [cell.value for cell in master_sheet['Q']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        MINDANAO_SPREMIUM_EAN = column_values[3:last_row_master]

        ##############################################
        ### HYBRID
        ##############################################

        #GMA 
        column_values = [cell.value for cell in master_sheet['T']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        GMA_HYBRID_EAN = column_values[3:last_row_master]

        #LUZON 
        column_values = [cell.value for cell in master_sheet['U']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        LUZON_HYBRID_EAN = column_values[3:last_row_master]

        #VISAYAS
        column_values = [cell.value for cell in master_sheet['V']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        VISAYAS_HYBRID_EAN = column_values[3:last_row_master]

        #MINDANAO
        column_values = [cell.value for cell in master_sheet['W']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        MINDANAO_HYBRID_EAN = column_values[3:last_row_master]

        ##############################################
        ### HEAVY RESELLER
        ##############################################

        #GMA 
        column_values = [cell.value for cell in master_sheet['Z']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        GMA_HR_EAN = column_values[3:last_row_master]

        #LUZON 
        column_values = [cell.value for cell in master_sheet['AA']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        LUZON_HR_EAN = column_values[3:last_row_master]

        #VISAYAS
        column_values = [cell.value for cell in master_sheet['AB']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        VISAYAS_HR_EAN = column_values[3:last_row_master]

        #MINDANAO
        column_values = [cell.value for cell in master_sheet['AC']]
        last_row_master = max(index for index, value in enumerate(column_values, start=1) if value is not None)
        MINDANAO_HR_EAN = column_values[3:last_row_master]


        ##### populating store types with EANs

        # Load the master workbook
        master_wb = openpyxl.load_workbook(file2)

        # Specify the sheet names
        output_sheet = master_wb['Sheet2']

        ##############################
        ##             GMA MASS        
        ##############################
        # Values to update (adjust as needed)
        values_to_update = GMA_MASS_EAN
        storeType = GMA_MASS

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=1) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=2)  # Adjust the column index as needed
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]

        ##############################
        ##            LUZON MASS        
        ##############################

        values_to_update = LUZON_MASS_EAN
        storeType = LUZON_MASS

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=4) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=5)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]

        ##############################
        ##            VISAYAS MASS        
        ##############################

        values_to_update = VISAYAS_MASS_EAN
        storeType = VISAYAS_MASS

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=7) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=8)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]


        ##############################
        ##            MINDANAO MASS        
        ##############################

        values_to_update = MINDANAO_MASS_EAN
        storeType = MINDANAO_MASS

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=10) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=11)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]

        ##############################
        ##            GMA PREMIUM        
        ##############################

        values_to_update =  GMA_PREMIUM_EAN
        storeType = GMA_P

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=13) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=14)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]


        ##############################
        ##            LUZON PREMIUM        
        ##############################
        values_to_update =  LUZON_PREMIUM_EAN
        storeType = LUZON_P

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=16) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=17)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]

        ##############################
        ##            VISAYAS PREMIUM        
        ##############################
        values_to_update = VISAYAS_PREMIUM_EAN
        storeType = VISAYAS_P

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=19) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=20)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]

        ##############################
        ##           MINDANAO PREMIUM        
        ##############################
        values_to_update = MINDANAO_PREMIUM_EAN
        storeType = MINDANAO_P

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=22) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=23)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]

        ##############################
        ##          GMA SPREMIUM        
        ##############################
        values_to_update = GMA_SPREMIUM_EAN
        storeType = GMA_SP

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=25) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=26)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]

        ##############################
        ##          LUZON SPREMIUM        
        ##############################
        values_to_update = LUZON_SPREMIUM_EAN
        storeType = LUZON_SP

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=28) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=29)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]


        ##############################
        ##          VISAYAS SPREMIUM        
        ##############################
        values_to_update = VISAYAS_SPREMIUM_EAN
        storeType = VISAYAS_SP

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=31) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=32)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]

        ##############################
        ##         MINDANAO SPREMIUM        
        ##############################
        values_to_update = MINDANAO_SPREMIUM_EAN
        storeType = MINDANAO_SP

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=34) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=35)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]

        ##############################
        ##         GMA HYBRID     
        ##############################
        values_to_update = GMA_HYBRID_EAN
        storeType = GMA_H

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=37) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=38)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]

        ##############################
        ##         LUZON HYBRID     
        ##############################
        values_to_update = LUZON_HYBRID_EAN
        storeType = LUZON_H

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=40) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=41)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]

        ##############################
        ##         VISAYAS HYBRID     
        ##############################
        values_to_update = VISAYAS_HYBRID_EAN
        storeType = VISAYAS_H

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=43) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=44)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]

        ##############################
        ##      MINDANAO HYBRID     
        ##############################
        values_to_update = MINDANAO_HYBRID_EAN
        storeType = MINDANAO_H

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=46) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=47)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]

        ##############################
        ##      GMA HEAVY RESELLER  
        ##############################
        values_to_update = GMA_HR_EAN
        storeType = GMA_HR

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=49) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=50)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]

        ##############################
        ##      LUZON HEAVY RESELLER  
        ##############################
        values_to_update = LUZON_HR_EAN
        storeType = LUZON_HR

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=52) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=53)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]


        ##############################
        ##      VISAYAS HEAVY RESELLER  
        ##############################
        values_to_update = VISAYAS_HR_EAN
        storeType = VISAYAS_HR

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=55) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=56)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]

        ##############################
        ##      MINDANAO HEAVY RESELLER  
        ##############################
        values_to_update = MINDANAO_HR_EAN
        storeType = MINDANAO_HR

        idx = 0
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=58) # Adjust the column index as needed
            cell_to_update.value = storeType[idx-1]
            if x % len(values_to_update) == 0:
                idx += 1

        y = 0
        idx = 1
        for x in range(1,len(values_to_update)*len(storeType)+1):
            cell_to_update = output_sheet.cell(row=x + 1, column=59)  # Adjust the column index as needed
            print(x)
            if x % len(values_to_update) == 0:
                y = len(values_to_update)*idx
                idx += 1
            cell_to_update.value = values_to_update[x-1 -y]



        # Save the changes to the master workbook
        master_wb.save(file2)

        # Close the workbook
        master_wb.close()

        self.close_loading_screen()



    def show_loading_screen(self):
        self.loading_window = tk.Toplevel(self.master)
        self.loading_window.title("Loading...")

        # Label to display progress text
        progress_label = tk.Label(self.loading_window, text="Reading all files: 0%", padx=20, pady=20)
        progress_label.pack()

        # Progress bar
        progress_bar = ttk.Progressbar(self.loading_window, variable=self.progress_var, maximum=100, length=200, mode='determinate')
        progress_bar.pack(padx=20, pady=10)

        self.master.update()

        # Store progress label and bar in instance variables
        self.progress_label = progress_label
        self.progress_bar = progress_bar


    def update_progress(self, value):
        self.progress_var.set(value)
        self.loading_window.update()

        # Update progress label text
        self.progress_label.config(text=f"Loading: {int(value)}%")

    def close_loading_screen(self):
        if self.loading_window:
            self.loading_window.destroy()
            self.loading_window = None

    def process_files(self, file1):
        
        self.update_progress(0)
        self.close_loading_screen()

if __name__ == "__main__":
    root = tk.Tk()
    app = FileSelectorApp(root)
    root.mainloop()
